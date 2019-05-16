import openpyxl
import os
# os.chdir 是 python 切換到電腦指定路徑的方法
os.chdir(os.getcwd())
# 開啓名爲 “TW2330” 的 excel 檔案，存入 workbook 這個變數中
workbook = openpyxl.load_workbook('test.xlsx')
# 從 workbook 中開啓一個名爲 'HelloWorld' 的工作表，存入 sheet 變數
sheetname = workbook['test']
# 把 “Hello World!” 這個字串賦值到 row 為 1, column 為 1 的,也就是儲存格 A1
sheetname.cell(row=1, column=4).value = "Hello World!"
for i in range(3, 100):
    # 截取 column 值為 2，也就是欄為 B 的儲存格
    price_today = sheetname.cell(row=i, column=2).value
    if (price_today==None):
        price_today=1
    # 截取 column 值為 2 (也就是欄為 B 的儲存格) 但是上一列，也就是昨天的收盤價
    price_yesterday = sheetname.cell(row=i-1, column=2).value
    if (price_yesterday==None):
        price_yesterday=1
    # 截取出兩者後，就可以算出該筆資料的日報酬率了
    daily_return = (price_today-price_yesterday)/price_yesterday
    # 最後再將算出的日報酬率寫入同一列，但是 column 值為 3 的儲存格 (也就是欄為 C 的儲存格)
    sheetname.cell(row=i, column=3).value = daily_return

# 存檔
workbook.save('test.xlsx')
